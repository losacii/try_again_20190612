from openpyxl import Workbook
import time

def task():
    from openpyxl import load_workbook
    from os import system as command
    
    # 载入总表和分表
    wbx = load_workbook('xx.xlsx')
    wby = load_workbook('Y.xlsx')

    # 各选择 第1张 sheet
    wsx = wbx[wbx.sheetnames[0]]
    wsy = wby[wby.sheetnames[1]]
    print("写入测试：10 人。。。")

    # 特定单元格数值
    for r in range(2, 841):
        lineValues = []
        for c in [chr(i) for i in range(66, 77)]:
            ind = "{}{}".format(c,r)
            lineValues.append(wsx[ind].value)
        name, gender, pos, eth, birth, idn, dang, edu, q, phone, addr = lineValues
        wsy['A2'].value = "编号：{}".format(r-1)
        wsy['B5'].value = name
        wsy['D5'].value = gender
        wsy['F5'].value = eth
        wsy['F3'].value = pos

        wsy['B9'].value = idn
        wsy['B6'].value = birth
        wsy['D6'].value = dang
        wsy['F6'].value = edu

        wsy['C17'].value = phone
        wsy['C18'].value = addr
        wsy['E9'].value = addr

        wsy['C12'].value = q

        wby.save("RES/{}.{}.xlsx".format(r-1, name,))
        print(r-1)
    print("完成！")





def tmp():
    pass
    fp = open("dates.txt")
    fx = open("res.txt", "w")
    for line in fp.readlines():
        line = "{}.{}.{}".format(line[:4], line[4:6], line[6:])
        fx.write(line)
    fp.close()
    fx.close()


if __name__ == "__main__":
    task()
    #tmp()