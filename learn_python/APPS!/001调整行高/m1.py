''' 批量调整行高: V1 '''
from openpyxl import load_workbook

fp = "target.xlsx"
START_LINE = 4
END_LINE = 59

wb = load_workbook(fp); print(wb.sheetnames)
ws = wb[wb.sheetnames[0]]

print(ws['C5'].value)

for i in range(START_LINE, END_LINE + 1):
    ws.row_dimensions[i].height += 26

wb.save("res.xlsx")
print(__doc__.center(80, '~'))