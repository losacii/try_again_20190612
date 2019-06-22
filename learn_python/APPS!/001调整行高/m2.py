# coding:utf-8
from os import system as command
from openpyxl import load_workbook
import sys
put = sys.stdout.write

fp = "data1.xlsx"

wb = load_workbook(fp); print(wb.sheetnames)
ws = wb[wb.sheetnames[0]]

cell_range = ws['A1':'C2']
for row in cell_range:
    li = [str(cell.value) for cell in row]
    txt = '\t'.join(li)
    print(txt)


# wb.save("res.xlsx")
# os.system("explorer .")
