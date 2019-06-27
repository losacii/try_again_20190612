# coding: utf-8
from openpyxl import load_workbook

wb = load_workbook("./data.xlsx"); print(wb.sheetnames)
#ws0 = wb[wb.sheetnames[0]]
ws1 = wb[wb.sheetnames[1]]
#ws2 = wb[wb.sheetnames[2]]
#ws3 = wb[wb.sheetnames[3]]
#ws4 = wb[wb.sheetnames[4]]



setvalue = ""
zhen = []
zhens = []
for row in range(3, 16):
    values = [ c.value for c in ws1[row][2:10] ]
    if values[0] != setvalue and not setvalue:
        print("空，添加！")
        setvalue = values[0]
        zhen.append(values)
    elif values[0] == setvalue:
        zhen.append(values)
        print("非空，添加！")
    else:
        print("Clear")
        print(zhen)
        zhens.append(zhen)
        zhen.clear()
        break
print(zhens)
        





'''
for row in range(3, 10):
    get_zhen = ws1["C{}".format(row)].value
    if get_zhen != zhen and bgroup is not None:
        x = ws1[row][3:10]
        for cell in x:
            print(cell.value)
'''
        





'''
for cell in lines:
    countline = 0
    for datacell in ws0['B']:
        countline += 1
        if str(cell) == str(datacell.value):
            li = [str(i.value) for i in ws0[countline]]
            # ws2.append(li) 
            print(li)
            break
'''

# li = [i.value for i in ws[2]] # 一行数值列表

# wb.save("res.xlsx")
# print("Done!")

