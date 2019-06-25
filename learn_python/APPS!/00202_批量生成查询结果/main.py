# coding: utf-8
from openpyxl import load_workbook
from mods import getLines

wb = load_workbook("./data.xlsx"); print(wb.sheetnames)
ws0 = wb[wb.sheetnames[0]]
ws1 = wb[wb.sheetnames[1]]
ws2 = wb[wb.sheetnames[2]]

lines = getLines(" Please Enter Search Targets: ".center(80, '~'))


for cell in lines:
    countline = 0
    for datacell in ws0['B']:
        countline += 1
        if str(cell) == str(datacell.value):
            li = [str(i.value) for i in ws0[countline]]
            # ws2.append(li) 
            print(li)
            break

# li = [i.value for i in ws[2]] # 一行数值列表

# wb.save("res.xlsx")
# print("Done!")

