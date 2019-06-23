# coding: utf-8
from openpyxl import load_workbook

wb = load_workbook("./data.xlsx"); print(wb.sheetnames)
ws0 = wb[wb.sheetnames[0]]
ws1 = wb[wb.sheetnames[1]]
ws2 = wb[wb.sheetnames[2]]

print(str(ws0['B2'].value) == str(ws1['A6'].value))

for cell in ws1['A']:
    countline = 0
    for datacell in ws0['B']:
        countline += 1
        if str(cell.value) == str(datacell.value):
            li = [str(i.value) for i in ws0[countline]]
            ws2.append(li) #print(li)
            break

# li = [i.value for i in ws[2]] # 一行数值列表

wb.save("res.xlsx")
print("Done!")
