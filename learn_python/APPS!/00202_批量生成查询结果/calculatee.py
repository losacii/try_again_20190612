
# coding: utf-8
from openpyxl import load_workbook
from mods import showli

wb = load_workbook("./data.xlsx"); print(wb.sheetnames)
ws = wb[wb.sheetnames[0]]

x = ws['C3'].value
print(x)

setyuan = ""
agroup = []
bgroup = []

print(agroup.__len__())

for r in range(2, 25):
    values = [ cell.value for cell in ws[r] ]
    yuan = values[3]
    if len(agroup) == 0:
        setyuan = yuan; # print('empty person list')
        agroup.append(values)
    elif setyuan == yuan:
        agroup.append(values); # print('same not empty list')
    else:
        # showli(agroup)
        bgroup.append(agroup.copy()); # print('----> list not empty, but new name')
        agroup.clear()
        setyuan = yuan
        agroup.append(values)

bgroup.append(agroup)

count = 0
for group in bgroup:
    count += 1
    print('-----------> group {}:'.format(count))
    showli(group)

names = [cell.value for cell in ws[1]]
v1 = [cell.value for cell in ws[2]]
v2 = [cell.value for cell in ws[3]]

print(dict(zip(names, v1)))
print(dict(zip(names, v2)))
print(list(zip(names, v2)))
print(tuple(zip(names, v2)))


# z = zip(ws[1], ws[2], ws[3], ws[4])
# print(list(z))
# print(ws[1].values)