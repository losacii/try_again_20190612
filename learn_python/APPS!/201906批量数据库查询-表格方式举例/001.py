# coding:utf-8
from os import system as cmd
from openpyxl import Workbook

wb = Workbook() # 创建: 工作簿
ws = wb.active  # 选择: 活动表

# 写入: 两种方式
ws['B1'] = "B1 value!"
ws.append([3, 5, 9])
ws.append([1, 2, 8, 27])
ws.append([1, 2, 8, 27])

# 保存， 打开文件夹
wb.save("res.xlsx")
cmd("explorer .")


'''
    # 创建工作簿, 选择活动表
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

    # 设置表名称
    ws.title = "New Title"

    # 表标签的颜色
    ws.sheet_properties.tabColor = "1072BA"

'''