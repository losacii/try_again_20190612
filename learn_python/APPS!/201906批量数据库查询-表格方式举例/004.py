
# coding:utf-8
import os
from os import system as cmd
from openpyxl import load_workbook

def adjust(fp, start, end):
    wb = load_workbook(fp)
    ws = wb[wb.sheetnames[0]]

    for i in range(start, end + 1):
        h = ws.row_dimensions[i].height
        if h:
            h += 25
            #print(i, "add")
        else:
            h = 45
            #print(i, 45)

    wb.save(fp)

workDir = r"G:\USERS\tx1\work_DIR\TMP_work\files"
for f in os.listdir(workDir):
    fp = os.path.join(workDir, f)
    print("\n--------> ", f)
    adjust(fp, 4, 59)
    print("Done!", fp)



'''
wb = load_workbook("data.xlsx"); print(wb.sheetnames)
ws = wb[wb.sheetnames[0]]

ws.row_dimensions[3].height += 45

for cell in ws['2']:
    print(cell.value, end=',')

#wb.save("res.xlsx")
#cmd("explorer .")
'''