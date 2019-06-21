# coding:utf-8
from os import system as cmd
from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet("my sheet")

ws['B1'] = "B1 value!"
ws.append([3, 5, 9])
ws.append([1, 2, 8, 27])
ws.append([1, 2, 8, 27])

wb.save("mysheet.xlsx")
cmd("explorer .")

'''
创建表: 
    You can create new worksheets using the Workbook.create_sheet() method:
        ws1 = wb.create_sheet("Mysheet") # insert at the end (default)
        ws2 = wb.create_sheet("Mysheet", 0) # insert at first position
所有表名称
    wb.sheetnames

表与工作簿的包含关系: in
    for sheet in wb:
        print(sheet.title)

单元格， 取值、赋值
    c = ws['A4']
    ws['A4'] = 17

    d = ws.cell(row=4, column=2, value=10)

行高
    ws.row_dimensions[3].height += 35    



'''
