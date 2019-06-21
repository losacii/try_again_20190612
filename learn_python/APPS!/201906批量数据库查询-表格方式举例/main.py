# coding: utf-8
from os import system as cmd
from openpyxl import Workbook

wb = Workbook()    #创建文件对象

# grab the active worksheet
ws = wb.active     #获取第一个sheet

# Data can be assigned directly to cells
ws['A1'] = 42      #写入数字
ws['B1'] = "你好"+"automation test" #写入中文（unicode中文也可）

# Rows can also be appended
ws.append([1, 2, 3])    #写入多个单元格
ws.append([9, 2, 3])    #写入多个单元格

# Save the file
wb.save("./sample.xlsx")

cmd("explorer .")