# coding:utf-8
from os import system as cmd
from openpyxl import load_workbook

wb = load_workbook("data.xlsx"); print(wb.sheetnames)
ws = wb[wb.sheetnames[0]]

ws.row_dimensions[3].height += 45

for cell in ws['2']:
    print(cell.value, end=',')

#wb.save("res.xlsx")
#cmd("explorer .")


'''
    # 创建工作簿, 选择活动表
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

    # 设置表名称
    ws.title = "New Title"

    # 表标签的颜色
    ws.sheet_properties.tabColor = "1072BA"

'''